"""
This utility generates the pen testing's assessment activity tracker.
It takes json input and generates an excel worksheet
"""
# Risk & Vulnerability Assessment Reporting Engine

# Copyright 2022 The Risk & Vulnerability Reporting Engine Contributors, All Rights Reserved.
# (see Contributors.txt for a full list of Contributors)

# SPDX-License-Identifier: BSD-3-Clause

# Please see additional acknowledgments (including references to third party source code, object code, documentation and other files) in the license.txt file or contact permission@sei.cmu.edu for full terms.

# Created, in part, with funding and support from the United States Government. (see Acknowledgments file).

# DM22-1011
import sys
import os.path
import argparse
import datetime
import openpyxl as px
from openpyxl.styles import Alignment, NamedStyle, Font

import report_gen.utilities.assessment_facts as af

# ---- openpyxl settings

CENTER = Alignment(horizontal='center', vertical='center')
BOLD14 = px.styles.Font(bold=True, size=14)
BOLD = px.styles.Font(bold=True)
ITALIC = px.styles.Font(italic=True)
DOUBLE = px.styles.Side(border_style="hair")
TOP = px.styles.Border(top=DOUBLE)
BOTTOM = px.styles.Border(bottom=DOUBLE)


def write_header(ws, lst, r, border=True, space=True, cap=True):
    """
    lst - contains the column headers
    r - is the starting row_dimensions
    border - add a border to the bottom
    space - size the columns approx to size of string
    cap - capitilize name via json_name function
    note: no column is needed it since it is based on size of list
    """
    for col, name in enumerate(lst):
        # set style for header row
        if cap:
            display_name = json_name(name)
        else:
            display_name = name
        ws.cell(row=r, column=col + 1).value = display_name
        ws.cell(row=r, column=col + 1).font = BOLD14
        ws.cell(row=r, column=col + 1).alignment = CENTER
        if border:
            ws.cell(row=r, column=col + 1).border = BOTTOM

        # adjust the column widths
        if space:
            col_ltr = px.utils.get_column_letter(col + 1)
            ws.column_dimensions[col_ltr].width = 30


# ---- tracker parameters

tabs = [
    "infrastructure",
    "lateralmovement",
    "files",
    "interactivelogons",
    "highimpactscans",
    "significantevents",
    "artifact"
]

tabs_display_name = {
    "infrastructure": "Infrastructure",
    "lateralmovement": "Lateral Movement",
    "files": "Files",
    "interactivelogons": "Interactive Logons",
    "highimpactscans": "High Impact Scans",
    "significantevents": "Significant Events",
    "artifact": "Artifacts"
}

cols = {}

cols["infrats"] = ["hostname", "ip_address", "domain", "beacon_kill_date"]
cols["infraphishing"] = ["hostname", "ip_address", "domain"]
cols["infraredirectors"] = ["url"]
cols["infraws"] = ["hostname", "ip_address", "operating_system", "operator"]

cols["lateralmovement"] = [
    "initial_beacon",
    "hostname",
    "ip_address",
    "account_used",
    "host_moved_from",
    "movement_method",
    "callback_server",
    "notes",
]

cols["files"] = [
    "hostname",
    "ip_address",
    "file_location",
    "file_name",
    "status",
    "datetime_created",
    "datetime_deleted",
]

cols["interactivelogons"] = [
    "hostname",
    "ip_address",
    "account",
    "method",
    "logon_datetime",
    "logoff_datetime",
    "operator",
    "notes",
]

cols["highimpactscans"] = [
    "scan_type",
    "tool_used",
    "ip_ranges_targeted",
    "domains_targeted",
    "scan_start",
    "scan_end",
    "notes",
]

cols["significantevents"] = [
    "event",
    "notes",
    "start_datetime",
    "end_datetime",
]

cols["artifact"] = [
    "file_name",
    "description",
    "md5",
    "sha1",
    "sha256",
]


def json_name(s):
    """Takes django/json name and converts to tab ready name"""
    lst = s.split('_')
    lst = [x.capitalize() for x in lst]
    # special fixup for ip -> IP
    if "Ip" in lst:
        idx = lst.index('Ip')
        lst[idx] = lst[idx].upper()
    # special fixup for url -> URL
    if "Url" in lst:
        idx = lst.index('Url')
        lst[idx] = lst[idx].upper()
    # special fixup for md5 -> MD5
    if "Md5" in lst:
        idx = lst.index('Md5')
        lst[idx] = lst[idx].upper()
    # special fixup for sha1 -> SHA1
    if "Sha1" in lst:
        idx = lst.index('Sha1')
        lst[idx] = lst[idx].upper()
    # special fixup for sha256 -> SHA256
    if "Sha256" in lst:
        idx = lst.index('Sha256')
        lst[idx] = lst[idx].upper()
    # special fixup for date/time
    if "Datetime" in lst:
        idx = lst.index("Datetime")
        lst[idx] = "Date/Time"

    return ' '.join(lst)


def add_infra_details(ws, tracker_data):

    # find the teamserver information
    ts_fnds = []
    for cnt, item in enumerate(af.model_gen(tracker_data, 'ptportal.infrats')):
        ele = item['fields']
        ts_fnds.append(ele)

    curr_row = 1
    ws.merge_cells('A' + str(curr_row) + ':D' + str(curr_row))
    write_header(
        ws, ["TEAMSERVERS"], curr_row, border=False, space=False, cap=False
    )

    curr_row += 1

    add_sheet_details(ws, cols["infrats"], ts_fnds, curr_row)
    if isinstance(ts_fnds, str):  # didn't find a dictionary with data
        no_of_rows = 0
    else:
        no_of_rows = len(ts_fnds)
    curr_row += no_of_rows + 3  # add some blank rows

    # find the phishing information
    ph_fnds = []
    for cnt, item in enumerate(af.model_gen(tracker_data, 'ptportal.infraphishing')):
        ele = item['fields']
        ph_fnds.append(ele)

    ws.merge_cells('A' + str(curr_row) + ':D' + str(curr_row))
    write_header(
        ws, ["PHISHING"], curr_row, border=False, space=False, cap=False
    )

    curr_row += 1

    add_sheet_details(ws, cols["infraphishing"], ph_fnds, curr_row)
    if isinstance(ph_fnds, str):  # didn't find a dictionary with data
        no_of_rows = 0
    else:
        no_of_rows = len(ph_fnds)
    curr_row += no_of_rows + 3  # add some blank rows

    # find the redirector information
    rd_fnds = []
    for cnt, item in enumerate(af.model_gen(tracker_data, 'ptportal.infraredirectors')):
        ele = item['fields']
        rd_fnds.append(ele)

    ws.merge_cells('A' + str(curr_row) + ':D' + str(curr_row))
    write_header(
        ws, ["REDIRECTORS"], curr_row, border=False, space=False, cap=False
    )

    curr_row += 1

    add_sheet_details(ws, cols["infraredirectors"], rd_fnds, curr_row)
    if isinstance(rd_fnds, str):  # didn't find a dictionary with data
        no_of_rows = 0
    else:
        no_of_rows = len(rd_fnds)
    curr_row += no_of_rows + 3  # add some blank rows

    # find the workstation information
    ws_fnds = []
    for cnt, item in enumerate(af.model_gen(tracker_data, 'ptportal.infraws')):
        ele = item['fields']
        ws_fnds.append(ele)
    print(ws_fnds)

    # add a title
    ws.merge_cells('A' + str(curr_row) + ':D' + str(curr_row))
    write_header(
        ws, ["WORKSTATIONS"], curr_row, border=False, space=False, cap=False
    )

    curr_row += 1

    add_sheet_details(ws, cols["infraws"], ws_fnds, curr_row)


def add_sheet_details(ws, col_names, fnds, start_row=1):
    # write the header row
    for c, name in enumerate(col_names):
        # set style for header row
        write_header(ws, col_names, start_row, border=False, space=True)

    # add each finding to subsequent rows
    if isinstance(fnds, str):  # didn't find a dictionary with data
        return

    for r, ele in enumerate(fnds):
        for c, name in enumerate(col_names):
            ws.cell(row=start_row + r + 1, column=c + 1).value = ele[name]


def create_tracker(outfile, json_file):

    tracker_data = af.load_rva_info(json_file)
    # create the excel spreadsheet
    wb = px.Workbook()

    # name the initial sheet (deleted later if no coverpage wanted)
    ws = wb.active
    ws.title = "Cover Page"

    # create all the other tabs
    for t in tabs:
        ws = wb.create_sheet(tabs_display_name[t])

    # create the special case infra sheet
    add_infra_details(wb["Infrastructure"], tracker_data)

    # populate the rest of the tabs
    for t in tabs[1:]:
        if t in cols:
            fnds = []
            for cnt, item in enumerate(af.model_gen(tracker_data, 'ptportal.' + t)):
                ele = item['fields']
                fnds.append(ele)
            add_sheet_details(wb[tabs_display_name[t]], cols[t], fnds)

    # clean up and save out spreadsheet
    del wb["Cover Page"]
    wb.save(outfile)


def main():
    description = "generate Assessment Activity Tracker"
    parser = argparse.ArgumentParser(description=description)

    parser.add_argument(
        "-o",
        "--output_file",
        action="store",
        default="Assessment Activity Tracker.xlsx",
        help="Report file name",
    )

    parser.add_argument("-j", "--json_file", action="store", required=True)

    args = parser.parse_args()

    create_tracker(args.output_file, args.json_file)


if __name__ == '__main__':
    main()
