# Risk & Vulnerability Assessment Reporting Engine

# Copyright 2022 The Risk & Vulnerability Reporting Engine Contributors, All Rights Reserved.
# (see Contributors.txt for a full list of Contributors)

# SPDX-License-Identifier: BSD-3-Clause

# Please see additional acknowledgments (including references to third party source code, object code, documentation and other files) in the license.txt file or contact permission@sei.cmu.edu for full terms.

# Created, in part, with funding and support from the United States Government. (see Acknowledgments file).

# DM22-1011

import sys, os.path, argparse, datetime
import report_gen.utilities.assessment_facts as af

import openpyxl as xl
from bs4 import BeautifulSoup

"""
Mitigation Action
TBD, Fully Mitigated, Partially Mitigated, False Positive, Accepted Risk, No Action Taken

Challenges
No Issues - Recommended mitigations were sufficient, Cybersecurity/IT Budget Limitations, Staffing Limitations, Timing/Schedule Limitation, N/A - Researched and deemed False Positive, Technically INfeasible, Resource Constraints, Other- Please Explain
"""


def insert_findings(workbook, data):
    ws = workbook['2. Mitigated Findings']
    findings = af.model_gen(data, 'ptportal.uploadedfinding')
    thin = xl.styles.Side(border_style="thin", color="000000")
    border_style = xl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

    for index, finding in enumerate(findings):
        ws['A' + str(index + 2)].value = finding['fields']['severity']
        ws['B' + str(index + 2)].value = finding['fields']['uploaded_finding_name']
        ws['C' + str(index + 2)].value = finding['fields']['assessment_type']

        ws['A' + str(index + 2)].border = border_style
        ws['B' + str(index + 2)].border = border_style
        ws['C' + str(index + 2)].border = border_style
        ws['D' + str(index + 2)].border = border_style
        ws['E' + str(index + 2)].border = border_style
        ws['F' + str(index + 2)].border = border_style
        ws['G' + str(index + 2)].border = border_style


def insert_assessinfo(workbook, data):
    ws = workbook['1. Assessment Information']
    assessment_type = af.get_db_info(data, 'report.fields.report_type', 'Report Type')

    ws['B3'].value = af.get_db_info(
        data, 'engagementmeta.fields.asmt_id', 'Assessment ID'
    )
    ws['B7'].value = af.get_db_info(
        data, 'engagementmeta.fields.ext_start_date', 'Start Date'
    )
    ws['B8'].value = af.get_db_info(
        data, 'engagementmeta.fields.int_end_date', 'End Date'
    )
    ws['A10'].value = BeautifulSoup(
        af.get_db_info(data, 'report.fields.cisa_results', 'Report Summary'),
        'html.parser',
    ).get_text('\n')

    if assessment_type == 'RVA':
        ws['B2'].value = "Risk and Vulnerability Assessment (RVA)"
    else:
        ws['B2'].value = "Remote Penetration Test (RPT)"


def generate_remediation(template, json_file, output="RVA_Remediation.xlsm"):
    wb = xl.load_workbook(filename=template, keep_vba=True)
    data = af.load_asmt_info(json_file)

    insert_assessinfo(wb, data)
    insert_findings(wb, data)
    wb.save(output)


def main():
    description = "Generate remediation follow up sheet for assessment."
    parser = argparse.ArgumentParser(description=description)

    parser.add_argument("TEMPLATE", help="Remediation Workbook Template.")
    parser.add_argument(
        "-o", "--output_file", action="store", default="RVA_Remediation.xlsm"
    )
    parser.add_argument("-j", "--json_file", action="store", required=True)
    args = parser.parse_args()

    if not os.path.exists(args.json_file):
        print("Invalid JSON file: ", args.json_file)
        sys.exit(1)
    if not os.path.exists(args.TEMPLATE):
        print("Invalid template file: ", args.TEMPLATE)
        sys.exit(1)

    generate_remediation(args.TEMPLATE, args.json_file, args.output_file)


if __name__ == '__main__':
    main()
